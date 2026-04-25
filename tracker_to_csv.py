#!/usr/bin/env python3
"""
tracker_to_csv.py — Great Lake Cleaners
Converts the outing tracker to cleanups.csv for WordPress import.

Supports two data sources — tried in this order:
  1. Google Sheets  (if config.toml is present and --no-sheets is not set)
  2. Local .xlsx    (fallback, or when --xlsx / positional arg is passed)

Usage:
    python tracker_to_csv.py                         # Google Sheets via config.toml
    python tracker_to_csv.py --xlsx tracker.xlsx     # local file override
    python tracker_to_csv.py --no-sheets             # skip Sheets, use default xlsx
    python tracker_to_csv.py -o path/to/out.csv      # custom output path
    python tracker_to_csv.py --config my.toml        # custom config file

config.toml format:
    spreadsheet_id   = "1fs2eG1SvjJjwHcHPdan5lzHB_kXXriON"
    credentials_file = "credentials.json"
    sheet_name       = "Daily Log"              # optional, default: "Daily Log"
    output           = "cleanups/cleanups.csv"  # optional

Dependencies:
    pip install openpyxl gspread google-auth
"""

import sys
import csv
import argparse
import tomllib
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ── Column indices (0-based, after skipping 2 header rows) ───────────────────
COL_DATE         = 0
COL_LOCATION     = 1
COL_DURATION_MIN = 2
COL_BAGS         = 3
COL_GARBAGE_KG   = 4
COL_NOTES        = 5
COL_CANS         = 6
COL_BOTTLES      = 7
COL_RECYCLED_KG  = 8   # weight of recyclables (cans + bottles) — tracked separately from debris
COL_VOLUNTEERS   = 9
COL_NOTABLE      = 10
COL_GPS_LAT      = 11
COL_GPS_LON      = 12
COL_INSTAGRAM    = 13
COL_CORRIDOR     = 14
COL_TIRES        = 15

DEFAULT_XLSX   = "Great_Lake_Cleaners_Outing_Tracker.xlsx"
DEFAULT_CONFIG = "config.toml"
DEFAULT_OUTPUT = "cleanups/cleanups.csv"
SHEET_NAME     = "Daily Log"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _float(val, default=0.0):
    try:
        return float(val) if val not in (None, "") else default
    except (TypeError, ValueError):
        return default

def _int(val, default=0):
    try:
        return int(float(val)) if val not in (None, "") else default
    except (TypeError, ValueError):
        return default

def _str(val):
    return str(val).strip() if val is not None else ""

def _cell(row, idx):
    return row[idx] if len(row) > idx else ""


# ── Config ────────────────────────────────────────────────────────────────────

def load_config(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    with open(config_path, "rb") as f:
        return tomllib.load(f)


# ── Data loading ──────────────────────────────────────────────────────────────

def load_from_sheets(spreadsheet_id: str, credentials_file: str,
                     sheet_name: str = SHEET_NAME) -> list:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        sys.exit("Run: pip install gspread google-auth")

    creds_path = Path(credentials_file)
    if not creds_path.exists():
        sys.exit(
            f"Credentials file not found: {creds_path}\n"
            "Download your service account JSON key from Google Cloud Console\n"
            "and add it alongside config.toml. Never commit it to version control."
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds  = Credentials.from_service_account_file(str(creds_path), scopes=scopes)
    client = gspread.authorize(creds)

    try:
        sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)
    except gspread.exceptions.SpreadsheetNotFound:
        sys.exit(
            f"Spreadsheet not found: {spreadsheet_id}\n"
            "Check the ID and that the service account has been granted Viewer access."
        )
    except gspread.exceptions.APIError as e:
        if "400" in str(e):
            sys.exit(
                "API error 400: This is likely an .xlsx file stored in Google Drive\n"
                "rather than a native Google Sheet.\n"
                "Fix: Open the file in Google Drive → File → Save as Google Sheets.\n"
                "Then update spreadsheet_id in config.toml with the new URL."
            )
        raise
    except gspread.exceptions.WorksheetNotFound:
        sys.exit(f"Tab '{sheet_name}' not found in spreadsheet.")

    return _parse_rows(sheet.get_all_values(), source="Google Sheets")


def load_from_xlsx(xlsx_path: Path) -> list:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Run: pip install openpyxl")

    if not xlsx_path.exists():
        sys.exit(f"Tracker file not found: {xlsx_path}")

    wb = load_workbook(str(xlsx_path), read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"No '{SHEET_NAME}' sheet found in {xlsx_path.name}")

    rows = list(wb[SHEET_NAME].iter_rows(values_only=True))
    return _parse_rows(rows, source=xlsx_path.name)


def _parse_rows(all_rows: list, source: str) -> list:
    """Parse rows from either Sheets (strings) or xlsx (mixed types) into outing dicts."""
    outings = []

    for row_idx, row in enumerate(all_rows):
        if row_idx < 2:
            continue  # skip two header rows

        date_raw = _cell(row, COL_DATE)
        if not date_raw:
            break  # end of data

        if isinstance(date_raw, datetime):
            date_str = date_raw.strftime("%Y-%m-%d")
        else:
            date_str = _str(date_raw)
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    date_str = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue

        outings.append({
            "date":         date_str,
            "location":     _str(_cell(row, COL_LOCATION)) or "Unknown Location",
            "duration_min": _float(_cell(row, COL_DURATION_MIN)),
            "bags":         _float(_cell(row, COL_BAGS)),
            "garbage_kg":   _float(_cell(row, COL_GARBAGE_KG)),
            "cans":         _int(_cell(row, COL_CANS)),
            "bottles":      _int(_cell(row, COL_BOTTLES)),
            "volunteers":   max(1, _int(_cell(row, COL_VOLUNTEERS), default=1)),
            "notes":        _str(_cell(row, COL_NOTES)),
            "notable":      _str(_cell(row, COL_NOTABLE)),
            "gps_lat":      _str(_cell(row, COL_GPS_LAT)),
            "gps_lon":      _str(_cell(row, COL_GPS_LON)),
            "recycled_kg":  _float(_cell(row, COL_RECYCLED_KG)),
            "instagram":    _str(_cell(row, COL_INSTAGRAM)),
            "corridor":     _str(_cell(row, COL_CORRIDOR)),
            "tires":        _int(_cell(row, COL_TIRES)),
        })

    return outings


# ── Merge outings → events ────────────────────────────────────────────────────

def merge_to_events(outings: list) -> list:
    groups = defaultdict(list)
    for o in outings:
        groups[(o["date"], o["location"])].append(o)

    events = []
    for (date, location), group in sorted(groups.items()):
        bags           = sum(o["bags"]       for o in group)
        garbage_kg     = sum(o["garbage_kg"] for o in group)
        cans           = sum(o["cans"]       for o in group)
        bottles        = sum(o["bottles"]    for o in group)
        items_recycled = cans + bottles
        person_minutes = sum(o["duration_min"] * o["volunteers"] for o in group)
        hours          = round(person_minutes / 60, 2)
        volunteers     = max(o["volunteers"] for o in group)
        instagram_url  = next((o["instagram"] for o in group if o["instagram"]), "")

        notables    = [o["notable"] for o in group if o["notable"]]
        notable_str = "; ".join(notables)
        notes_str   = "; ".join(o["notes"] for o in group if o["notes"])

        recycle_parts = []
        if cans:    recycle_parts.append(f"{cans} cans")
        if bottles: recycle_parts.append(f"{bottles} bottles")
        if recycle_parts:
            notable_str = (notable_str + "; Recyclables: " + ", ".join(recycle_parts)).lstrip("; ")

        tires        = sum(o["tires"]       for o in group)
        recycled_kg  = sum(o["recycled_kg"] for o in group)
        gps_lat  = next((o["gps_lat"]  for o in group if o["gps_lat"]),  "")
        gps_lon  = next((o["gps_lon"]  for o in group if o["gps_lon"]),  "")
        corridor = next((o["corridor"] for o in group if o["corridor"]), "")

        events.append({
            "date":                date,
            "site_name":           location,
            "corridor":            corridor,
            "gps_lat":             gps_lat,
            "gps_lon":             gps_lon,
            "volunteers":          volunteers,
            "hours":               hours,
            "bags":                int(bags) if bags == int(bags) else bags,
            "weight_kg":           garbage_kg,
            "items_recycled":      items_recycled,
            "tires_removed":       tires or "",
            "recycled_weight_kg":  recycled_kg or "",
            "species_planted":     "",
            "meters_bank_cleared": "",
            "notable_finds":       notable_str,
            "wildlife_obs":        "",
            "notes":               notes_str,
            "photo_folder":        "",
            "best_photo":          "",
            "instagram_url":       instagram_url,
        })

    return events


# ── CSV output ────────────────────────────────────────────────────────────────

CSV_COLUMNS = [
    "date", "site_name", "corridor", "gps_lat", "gps_lon",
    "volunteers", "hours", "bags", "weight_kg",
    "items_recycled", "tires_removed", "recycled_weight_kg", "species_planted", "meters_bank_cleared",
    "notable_finds", "wildlife_obs", "notes",
    "photo_folder", "best_photo", "instagram_url",
]

def write_csv(events: list, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(events)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export Great Lake Cleaners outing tracker to cleanups.csv"
    )
    parser.add_argument("tracker", nargs="?",
        help="Path to local .xlsx (overrides Google Sheets)")
    parser.add_argument("--xlsx", metavar="FILE",
        help="Explicit local .xlsx path (overrides Google Sheets)")
    parser.add_argument("--no-sheets", action="store_true",
        help="Skip Google Sheets, use local xlsx")
    parser.add_argument("--config", metavar="FILE", default=DEFAULT_CONFIG,
        help=f"Path to config.toml (default: {DEFAULT_CONFIG})")
    parser.add_argument("--output", "-o", metavar="FILE",
        help=f"Output CSV path (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    config      = load_config(Path(args.config))
    output_path = Path(args.output or config.get("output", DEFAULT_OUTPUT))
    sheet_name  = config.get("sheet_name", SHEET_NAME)

    xlsx_override = args.tracker or args.xlsx
    use_sheets    = (
        not xlsx_override
        and not args.no_sheets
        and "spreadsheet_id" in config
    )

    if use_sheets:
        sid = config["spreadsheet_id"]
        creds = config.get("credentials_file", "credentials.json")
        print(f"Reading from Google Sheets ({sid[:16]}…) ...")
        outings = load_from_sheets(sid, creds, sheet_name)
    else:
        xlsx_path = Path(xlsx_override or DEFAULT_XLSX)
        print(f"Reading {xlsx_path.name} ...")
        outings = load_from_xlsx(xlsx_path)

    print(f"  {len(outings)} outing row(s) found")

    events = merge_to_events(outings)
    print(f"  {len(events)} cleanup event(s) after merging same-day/same-location outings")

    write_csv(events, output_path)
    print(f"\nWritten to {output_path}")
    print("\nEvents exported:")
    for e in events:
        gps_str = f"  [{e['gps_lat']}, {e['gps_lon']}]" if e["gps_lat"] else "  [no GPS]"
        print(f"  {e['date']}  {e['site_name']:<25}  {e['bags']} bags  {e['weight_kg']} kg"
              f"  {e['hours']}h  {e['volunteers']} vol  {e['items_recycled']} recycled{gps_str}")

    print("\nNext step: WP Admin → Tools → Import Cleanups CSV → upload the file above.")
    print("Duplicate date+site entries are skipped automatically.")


if __name__ == "__main__":
    main()
